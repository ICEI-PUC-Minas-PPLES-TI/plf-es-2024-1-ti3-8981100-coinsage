from datetime import datetime
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Border, PatternFill, Side
from sqlalchemy.orm import Session

from src.models.db.currency_base_info import CurrencyBaseInfoModel
from src.models.db.first_stage_analysis import FirstStageAnalysisModel
from src.models.db.wallet_transaction import WalletTransaction
from src.services.analysis.first_stage.closing_price_service import PriceService


class WorkbookService:
    def __init__(self, session: Session):
        self.session = session
        self.price_service = PriceService(self.session)

    def create_workbook(self, headers: List[str]) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active

        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            worksheet.column_dimensions[cell.column_letter].width = 100 * 0.18  # Set the width in pixels
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(wrap_text=True, horizontal="center", vertical="center")

        return workbook

    def style_workbook(self, workbook: Workbook) -> Workbook:
        self.set_min_height(workbook)
        self.style_yes_no(workbook)

        # general info
        self.apply_generic_style(workbook, 1, 4, True, "ffc0c0")
        # weekly analysis > 10%
        self.apply_generic_style(workbook, 5, 7, False, "b0f283")
        # ema8 relations
        self.apply_generic_style(workbook, 8, 13, False, "fce746")
        # >10% style
        self.apply_percentage_style(workbook)

        return workbook

    def apply_percentage_style(self, workbook: Workbook) -> Workbook:
        for row in workbook.active.iter_rows(min_row=1, max_row=workbook.active.max_row):
            for cell in row:
                if cell.row == 1:
                    continue
                if cell.column == 6:
                    if cell.value == "N/A":
                        continue
                    if cell.value >= 10:
                        cell.font = cell.font.copy(color="226625", bold=True)
                    elif cell.value < 0:
                        cell.font = cell.font.copy(color="FF0000", bold=True)

        return workbook

    def format_workbook(self, workbook: Workbook) -> Workbook:
        self.format_currency(workbook)
        self.format_percentage(workbook)

        return workbook

    def format_currency(self, workbook: Workbook) -> Workbook:
        for col in ["D", "G", "H", "I", "J"]:
            for cell in workbook.active[col]:
                if cell.row == 1:
                    continue
                self.modify_cell_format(cell, "#,##0.00", 0)

    def format_percentage(self, workbook: Workbook) -> Workbook:
        for cell in workbook.active["F"]:
            if cell.row == 1:
                continue
            self.modify_cell_format(cell, "0.00", 0)

    def modify_cell_format(self, cell, formatter: str, times: int) -> None:
        cell.number_format = formatter
        if self.check_zero(cell.value, times) and times < 6:
            self.modify_cell_format(cell, f"{formatter}{0}", times + 1)

    def check_zero(self, value: str, rounder: int) -> bool:
        try:
            return round(float(value), rounder + 2) == 0
        except Exception:
            return False

    def style_yes_no(self, workbook: Workbook) -> Workbook:
        # all yes and no should be bold
        # yes should be green and no red
        for row in workbook.active.iter_rows(min_row=1, max_row=workbook.active.max_row):
            for cell in row:
                if cell.value == "SIM":
                    cell.font = cell.font.copy(color="226625", bold=True)
                elif cell.value == "NÃO":
                    cell.font = cell.font.copy(color="FF0000", bold=True)
        return workbook

    def set_min_height(self, workbook: Workbook) -> Workbook:
        for row in workbook.active.iter_rows(min_row=1, max_row=workbook.active.max_row):
            for cell in row:
                if cell.row == 1:
                    continue
                workbook.active.row_dimensions[cell.row].height = 70

    def apply_generic_style(
        self, workbook: Workbook, start_col: int, end_col: int, wrap_text: bool, color_hex: str
    ) -> Workbook:
        for col in workbook.active.iter_cols(min_col=start_col, max_col=end_col):
            for cell in col:
                if cell.col_idx == start_col and cell.row != 1 and wrap_text:
                    cell.alignment = cell.alignment.copy(wrap_text=wrap_text)
                # color fill
                redFill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                cell.fill = redFill
                # border
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                cell.border = thin_border
                # center alignment
                cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

    def fill_workbook(self, workbook: Workbook, headers: List[str], analysis_uuid: str) -> Workbook:
        worksheet = workbook.active
        data = (
            self.session.query(FirstStageAnalysisModel)
            .filter(FirstStageAnalysisModel.uuid_analysis == analysis_uuid)
            .join(CurrencyBaseInfoModel, FirstStageAnalysisModel.uuid_currency == CurrencyBaseInfoModel.uuid)
        )

        btc = data.filter(CurrencyBaseInfoModel.symbol == "BTC").first()
        if btc:
            data = data.filter(CurrencyBaseInfoModel.symbol != "BTC").all()
            data = [btc] + data  # btc is added at the beginning of the list
        else:
            data = data.all()

        header_to_model_attr = {
            "SETOR": lambda item: (
                self.price_service._get_sector_by_symbol(item.currency.uuid).title
                if self.price_service._get_sector_by_symbol(item.currency.uuid).title != "Unknown"
                else "N/A"
            ),
            "CRIPTOMOEDA": lambda item: item.currency.symbol if item.currency else "N/A",
            "RANKING": lambda item: item.ranking if item.ranking else "N/A",
            "VALOR MERCADO (US$ BILHÕES)": lambda item: item.market_cap / 1_000_000_000 if item.market_cap else "N/A",
            "DATA VALORIZ. SEMANAL > 10%": lambda item: (
                datetime.strftime(item.last_updated, "%d/%m/%Y") if item.last_updated else "N/A"
            ),
            "VALORIZ. NESTA DATA (%)": lambda item: (
                item.week_increase_percentage if item.week_increase_percentage else "N/A"
            ),
            "PREÇO NO MOMENTO (US$)": "current_price",
            "PREÇO SEMANAL FECHAMENTO (US$)": "closing_price",
            "PREÇO SEMANAL ABERTURA (US$)": "open_price",
            "EMA(8) SEMANAL": lambda item: item.ema8 if item.ema8 else "N/A",
            "PREÇO SEMANAL FECHAMENTO > EMA (8)": lambda item: "SIM" if item.ema8_less_close else "NÃO",
            "EMA (8) > PREÇO SEMANAL ABERTURA": lambda item: "SIM" if item.ema8_greater_open else "NÃO",
            "MÉDIAS MÓVEIS DIÁRIAS ALINHADAS": lambda item: "SIM" if item.ema_aligned else "NÃO",
            "DATA AUMENTO DE VOLUME (d)": lambda item: item.increase_volume_day if item.increase_volume_day else "N/A",
            "AUMENTO DE VOLUME (w)": lambda item: item.increase_volume if item.increase_volume else "N/A",
            "AUMENTO DE VOLUME": lambda item: item.increase_volume if item.increase_volume else "N/A",
            "VOLUME ATUAL": lambda item: item.today_volume if item.today_volume else "N/A",
            "VOLUME ANTES DO AUMENTO": lambda item: (
                item.volume_before_increase if item.volume_before_increase else "N/A"
            ),
            "VOLUME > 200%": lambda item: "SIM" if item.expressive_volume_increase else "NÃO",
            "SINAL DE COMPRA": lambda item: "SIM" if item.buying_signal else "NÃO",
        }

        for row_idx, item in enumerate(data, start=2):
            for header in headers:
                col_idx = headers.index(header) + 1
                model_attr = header_to_model_attr.get(header)

                value = (
                    model_attr(item)
                    if callable(model_attr)
                    else getattr(item, model_attr, "N/A") if model_attr is not None else "N/A"
                )
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        return workbook

    def create_wallet_workbook(self, headers: List[str]) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active

        for col, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col, value=header)
            worksheet.column_dimensions[cell.column_letter].width = 100 * 0.18
            cell.font = cell.font.copy(bold=True)
            cell.alignment = cell.alignment.copy(wrap_text=True, horizontal="center", vertical="center")

        return workbook

    def fill_wallet_workbook(self, workbook: Workbook, headers: List[str], user_id: int) -> Workbook:
        worksheet = workbook.active
        transactions = (
            self.session.query(WalletTransaction, CurrencyBaseInfoModel)
            .where(WalletTransaction.uuid_currency == CurrencyBaseInfoModel.uuid)
            .filter(WalletTransaction.user_id == user_id)
            .all()
        )

        header_to_model_attr = {
            "CRIPTOMOEDA": lambda item: item.symbol if item.symbol else "N/A",
            "QUANTIDADE": lambda item: item.quantity if item.quantity else "N/A",
            "VALOR (US$)": lambda item: item.amount if item.amount else "N/A",
            "DATA": lambda item: item.date.strftime("%d/%m/%Y") if item.date else "N/A",
            "PREÇO NA COMPRA (US$)": lambda item: item.price_on_purchase if item.price_on_purchase else "N/A",
        }

        for row_idx, item in enumerate(transactions, start=2):
            for header in headers:
                col_idx = headers.index(header) + 1
                model_attr = header_to_model_attr.get(header)
                val = None
                if header == "CRIPTOMOEDA":
                    val = model_attr(item[1])
                else:
                    val = model_attr(item[0])
                value = (
                    val
                    if callable(model_attr)
                    else getattr(item, model_attr, "N/A") if model_attr is not None else "N/A"
                )

                worksheet.cell(row=row_idx, column=col_idx, value=value)
        return workbook
